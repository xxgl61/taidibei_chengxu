
import os
os.environ['ENABLE_PJRT_COMPATIBILITY'] = '1'
os.environ['JAX_PLATFORMS'] = 'METAL,cpu'


import pandas as pd 
import numpy as np
import math
import jax  #全部jax库
import jax.numpy as jnp #JAX的NumPy接口
from jax import lax
from openpyxl import load_workbook #Excel写入用
import rasterio   #读取TIF文件用
import os #文件路径处理
import glob #文件搜索用


# GPU设备初始化

try:
    METAL_DEVICE = jax.devices('METAL')[0]    # 尝试获取Apple GPU设备并获取第一个设备
    print(f"✅ 成功识别Apple GPU：{METAL_DEVICE}")
    USE_GPU = True
except Exception as e:
    print(f"⚠️ GPU初始化失败：{str(e)}，将使用CPU运行")
    METAL_DEVICE = jax.devices('cpu')[0]    # 回退到CPU设备，硬算，性能较差，建议换成cuda
    USE_GPU = False


def is_numeric_string(s):
    """判断字符串是否可转为数字"""
    if isinstance(s, (int, float)):
        return True
    if not isinstance(s, str):
        return False
    # 清洗特殊字符
    s_clean = s.strip().replace(',', '.').replace(' ', '')
    if s_clean in ['', 'NaN', 'nan', 'None', 'none']:
        return False
    try:
        float(s_clean)
        return True
    except (ValueError, TypeError):
        return False

def extract_coords_from_df(df, sheet_name):     #匹配坐标（示例数据有部分算不出）
    
    print(f"\n【{sheet_name}】工作表列名：{df.columns.tolist()}")
    X_COL_CANDIDATES = ['x', 'x坐标', '横坐标', 'x_coordinate', '经度', 'x轴']
    Y_COL_CANDIDATES = ['y', 'y坐标', '纵坐标', 'y_coordinate', '纬度', 'y轴']
    x_col = None
    y_col = None

    
    # 匹配x列
    for col in df.columns:
        col_lower = str(col).strip().lower()
        if any(cand.lower() in col_lower for cand in X_COL_CANDIDATES):
            x_col = col
            print(f"✅ 【{sheet_name}】自动匹配x列：{col}")
            break
    # 匹配y列
    for col in df.columns:
        if col == x_col:
            continue
        col_lower = str(col).strip().lower()
        if any(cand.lower() in col_lower for cand in Y_COL_CANDIDATES):
            y_col = col
            print(f"✅ 【{sheet_name}】自动匹配y列：{col}")
            break

    #自动匹配失败时手动选择
    if not x_col or not y_col:
        print(f" 【{sheet_name}】自动匹配失败，需手动选择列")
        print("列名列表（带序号）：")
        for idx, col in enumerate(df.columns, 1):
            sample = df[col].head(3).tolist()
            print(f"  {idx}. {col} → 前3个样本：{sample}")
        
        # 手动选x列
        while not x_col:
            try:
                x_idx = int(input(f"请输入x坐标列的序号（1~{len(df.columns)}）：")) - 1
                x_col = df.columns[x_idx]
                print(f" 手动选择x列：{x_col}")
            except (ValueError, IndexError):
                print(f" 序号无效，请输入1~{len(df.columns)}的数字")
        
        # 手动选y列（避免重复）
        while not y_col or y_col == x_col:
            try:
                y_idx = int(input(f"请输入y坐标列的序号（1~{len(df.columns)}，不能与x列相同）：")) - 1
                y_col = df.columns[y_idx]
                if y_col == x_col:
                    print(f" 不能选x列（{x_col}），请重新选择")
                else:
                    print(f" 手动选择y列：{y_col}")
            except (ValueError, IndexError):
                print(f" 序号无效，请输入1~{len(df.columns)}的数字")

    
    df_temp = df[[x_col, y_col]].copy()
    # 清洗字符串格式
    for col in [x_col, y_col]:
        if df_temp[col].dtype == 'object':
            df_temp[col] = df_temp[col].astype(str).apply(
                lambda s: s.strip().replace(',', '.').replace(' ', '')
            )
        df_temp[col] = df_temp[col].replace(['', 'NaN', 'nan'], np.nan)
    
    # 过滤非数字和空值
    df_clean = df_temp.dropna()
    df_clean = df_clean[
        df_clean[x_col].apply(is_numeric_string) &
        df_clean[y_col].apply(is_numeric_string)
    ]

    if df_clean.empty:
        print(f" 【{sheet_name}】无有效坐标数据（清洗后为空）")
        return None, None

    # 转换为数值数组
    x_codes = df_clean[x_col].astype(float).to_numpy()
    y_codes = df_clean[y_col].astype(float).to_numpy()
    print(f" 【{sheet_name}】提取y坐标{len(y_codes)}个（范围：{y_codes.min():.2f}~{y_codes.max():.2f}）")
    return x_codes, y_codes


def read_tif_to_array(tif_path):
    """读取TIF文件，返回x/y坐标数组和高程矩阵"""
    with rasterio.open(tif_path) as src:
        elevation_data = src.read(1)  # 读取第一波段（高程数据）
        transform = src.transform  # 获取仿射变换参数
        print("高程数据形状:", elevation_data.shape)
        print("仿射变换参数:", transform)
        # 计算每个像素的地理坐标
        rows, cols = elevation_data.shape
        x_coords = np.arange(cols) * transform.a + transform.c  # 计算x坐标
        y_coords = np.arange(rows) * transform.e + transform.f  # 计算y坐标
        return x_coords, y_coords, elevation_data
def read_csv_list(csv_path):
    """读取高程数据CSV文件，返回x/y坐标数组和高程矩阵"""
    try:
        df = pd.read_csv(csv_path, index_col=0)
        # 提取x/y坐标
        x_coords = np.array([float(col) for col in df.columns])
        y_coords = np.array([float(idx) for idx in df.index])
        elevation_matrix = df.values.astype(np.float32)
        print(f"\n✅ 成功读取高程数据：")
        print(f"  - x坐标范围：{x_coords.min():.2f}~{x_coords.max():.2f}（共{len(x_coords)}个）")
        print(f"  - y坐标范围：{y_coords.min():.2f}~{y_coords.max():.2f}（共{len(y_coords)}个）")
        print(f"  - 高程矩阵形状：{elevation_matrix.shape}")
        return x_coords, y_coords, elevation_matrix
    except Exception as e:
        print(f"❌ 读取高程数据失败：{str(e)}")
        return None, None, None

def read_excel_list(excel_path):
    """读取Excel数据"""
    try:
        xl_file = pd.ExcelFile(excel_path, engine="openpyxl")
        sheet_names = xl_file.sheet_names
        sheetnumbers = len(sheet_names)
        print(f"\n===== 读取Excel文件：{excel_path} =====")
        print(f"包含工作表（共{sheetnumbers}个）：")
        for i, name in enumerate(sheet_names, 1):
            print(f"  {i}. {name}")
        
       
        sheet_imput = True
        while sheet_imput:
            choice = input(f"\n请选择要计算的工作表（全选按0，单个按序号）：")
            if not choice.isdigit():
                print("❌ 请输入数字！")
                continue
            choice = int(choice)
            if choice < 0 or choice > sheetnumbers:
                print(f"❌ 序号无效，需输入0~{sheetnumbers}！")
                continue
            sheet_imput = False

        all_results = {}
        # 处理选择
        if choice == 0:
            target_sheets = sheet_names
            print(f"\n✅ 选择全选，将处理所有{len(target_sheets)}个工作表")
        else:
            target_sheets = [sheet_names[choice-1]]
            print(f"\n✅ 选择单个工作表：{target_sheets[0]}")

        # 读取每个目标工作表
        for sheet_name in target_sheets:
            print(f"\n----- 处理工作表：{sheet_name} -----")
            # 特殊处理第5个工作表
            if sheet_names.index(sheet_name) + 1 == 5:
                df = pd.read_excel(xl_file, sheet_name=sheet_name, usecols=["x坐标/m", "y坐标/m"])
            else:
                df = pd.read_excel(xl_file, sheet_name=sheet_name)
            
            # 提取坐标（传入工作表名称）
            x_codes, y_codes = extract_coords_from_df(df, sheet_name=sheet_name)
            if x_codes is not None and y_codes is not None:
                high_matrix = df.values.astype(np.float32)
                all_results[sheet_name] = (x_codes, y_codes, high_matrix)
                print(f"✅ 【{sheet_name}】读取完成，存入结果集")
            else:
                print(f"❌ 【{sheet_name}】无有效坐标，跳过")
        
        print(f"\n✅ Excel读取完成，共获取{len(all_results)}个有效工作表的坐标")
        return all_results
    except Exception as e:
        print(f"❌ 读取Excel失败：{str(e)}")
        return {}


@jax.jit #编译成GPU加速函数 
def jax_conv2d(input_matrix, kernel):
    """JAX原生二维卷积（计算坡度坡向用）"""
    H, W = input_matrix.shape  #输入矩阵大小
    kH, kW = kernel.shape #获取卷积核大小
    pad_h = kH // 2
    pad_w = kW // 2
    
    padded = jnp.pad(input_matrix, ((pad_h, pad_h), (pad_w, pad_w)), mode="edge") #整理格式
    padded_4d = padded[None, None, :, :]
    kernel_4d = kernel[None, None, :, :]
    
    conv_result = lax.conv_general_dilated(
        lhs=padded_4d, rhs=kernel_4d, window_strides=(1,1), padding="VALID",
        dimension_numbers=("NCHW", "OIHW", "NCHW"), feature_group_count=1
    )
    return conv_result[0, 0, :, :]

@jax.jit
def calculate_slope_aspect_gpu(elevation_matrix, x_res, y_res):
    """全图坡度坡向计算（GPU加速）"""
    # Sobel卷积核
    kernel_x = jnp.array([[-1,0,1], [-2,0,2], [-1,0,1]]) / (8 * x_res)
    kernel_y = jnp.array([[-1,-2,-1], [0,0,0], [1,2,1]]) / (8 * y_res)
    
    dx = jax_conv2d(elevation_matrix, kernel_x)
    dy = jax_conv2d(elevation_matrix, kernel_y)
    
    slope = jnp.degrees(jnp.arctan(jnp.sqrt(dx**2 + dy**2)))
    aspect = 57.29578 * jnp.arctan2(dy, -dx)
    aspect = (aspect + 360) % 360
    return slope, aspect

@jax.jit
def find_nearest_indices(coords, targets):
    """增强版：批量匹配最近坐标（返回索引+距离，解决二级分水岭找不到的问题）"""
    distances_matrix = jnp.abs(coords[None, :] - targets[:, None])  # [M, N]
    indices = jnp.argmin(distances_matrix, axis=1)  # 最近索引
    distances = jnp.min(distances_matrix, axis=1)   # 最近距离
    return indices, distances

@jax.jit
def window_stats_gpu(elevation_matrix, y_indices, x_indices):
    """3x3窗口统计（GPU加速，无动态切片）"""
    padded_matrix = jnp.pad(elevation_matrix, ((1,1), (1,1)), mode="edge")
    pad_offset = 1
    window_size = 3

    def single_window(y_idx, x_idx):
        start_y = y_idx + pad_offset
        start_x = x_idx + pad_offset
        # 固定大小切片
        window = lax.dynamic_slice(padded_matrix, (start_y, start_x), (window_size, window_size))
        mask = ~jnp.isnan(window)
        valid_count = jnp.sum(mask)
        center_val = elevation_matrix[y_idx, x_idx]
        
        # 替换NaN为0，避免计算错误
        window_no_nan = jnp.where(mask, window, 0.0)
        sum_val = jnp.sum(window_no_nan)
        sum_sq_val = jnp.sum(window_no_nan ** 2)
        
        # 统计量计算
        mean_elev = jnp.where(valid_count >=5, sum_val/valid_count, center_val)
        max_elev = jnp.where(valid_count >=5, jnp.max(window_no_nan), center_val)
        min_elev = jnp.where(valid_count >=5, jnp.min(window_no_nan), center_val)
        variance = jnp.where(valid_count >=5, (sum_sq_val/valid_count)-(sum_val/valid_count)**2, 0.0)
        std_elev = jnp.sqrt(jnp.maximum(variance, 0.0))
        
        return max_elev, min_elev, mean_elev, std_elev
    
    return jax.vmap(single_window)(y_indices, x_indices)


def write_multi_sheet_excel(results_dict, output_path):
    """多工作表Excel写入（确保工作表可见）"""
    valid_results = {name: res for name, res in results_dict.items() if res}
    if not valid_results:
        print("\n❌ 无有效结果可写入Excel！")
        return

    try:
        # 写入Excel
        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            for sheet_name, results in valid_results.items():
                df = pd.DataFrame(results)
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                print(f"✅ 写入工作表「{sheet_name}」：{len(results)}条结果")
        
        # 确保第一个工作表可见
        wb = load_workbook(output_path)
        first_sheet = list(wb.sheetnames)[0]
        wb[first_sheet].sheet_state = "visible"
        wb.save(output_path)
        wb.close()
        
        print(f"\n✅ 所有结果已保存到：{output_path}")
    except Exception as e:
        print(f"\n❌ 写入Excel失败：{str(e)}")

        
        



def main():
    paper_needs="论文需要的数据.xlsx"
    ELEVATION_CSV = "陕甘八县的高程数据.csv"  # 附件1：高程数据
    EXCEL_PATH = "附件2  秦直道及周边地形和相关遗迹的数据.xlsx"  # 附件2：目标坐标
    TIF_PATH = "陕甘八县的高程数据.tif"  # 附件3：TIF格式高程数据
    OUTPUT_EXCEL = "地形特征计算结果_final2.xlsx"  # 输出结果
    
   
    # 读取高程数据
    x_coords_np, y_coords_np, e_matrix_np = read_tif_to_array(TIF_PATH)
    if x_coords_np is None or y_coords_np is None or e_matrix_np is None:
        print("\n❌ 高程数据读取失败，程序终止！")
        return
    
    # 读取Excel目标坐标
    all_results = read_excel_list(EXCEL_PATH)
    if not all_results:
        print("\n❌ 无有效工作表数据，程序终止！")
        return
    
    # 读取csv高程数据
    result= read_excel_list(EXCEL_PATH)
    if result is None:
        print("\n CSV高程数据读取失败，程序终止！")
        return  
    
    #读取论文要体现的数据
    paper=read_excel_list(paper_needs)
    if not all_results:
        print("\n❌ 无有效工作表数据，程序终止！")
        return
    
    

    # ----------------  数据加载到GPU ----------------
    if USE_GPU:
        print(f"\n===== 加载数据到GPU =====")
        x_coords_gpu = jax.device_put(jnp.array(x_coords_np), METAL_DEVICE)
        y_coords_gpu = jax.device_put(jnp.array(y_coords_np), METAL_DEVICE)
        e_matrix_gpu = jax.device_put(jnp.array(e_matrix_np), METAL_DEVICE)
        
        # 计算高程数据分辨率（x/y方向的平均间隔）
        x_res = float(jnp.mean(jnp.diff(x_coords_gpu))) if len(x_coords_gpu) > 1 else 1.0
        y_res = float(jnp.mean(jnp.diff(y_coords_gpu))) if len(y_coords_gpu) > 1 else 1.0
        print(f" 高程数据分辨率：x方向{x_res:.2f}m，y方向{y_res:.2f}m")
        
        # 一次性计算全图坡度坡向（GPU加速）
        print(" GPU计算全图坡度坡向...")
        slope_matrix_gpu, aspect_matrix_gpu = calculate_slope_aspect_gpu(e_matrix_gpu, x_res, y_res)
        print(" 全图坡度坡向计算完成")
    else:
        # CPU模式
        x_coords_gpu = x_coords_np
        y_coords_gpu = y_coords_np
        e_matrix_gpu = e_matrix_np
        x_res = np.mean(np.diff(x_coords_np)) if len(x_coords_np) > 1 else 1.0
        y_res = np.mean(np.diff(y_coords_np)) if len(y_coords_np) > 1 else 1.0
        
        # CPU计算坡度坡向（用scipy）
        from scipy.ndimage import convolve
        print("CPU计算全图坡度坡向...")
        kernel_x = np.array([[-1,0,1], [-2,0,2], [-1,0,1]])/(8*x_res)
        kernel_y = np.array([[-1,-2,-1], [0,0,0], [1,2,1]])/(8*y_res)
        dx = convolve(e_matrix_np, kernel_x, mode="nearest")
        dy = convolve(e_matrix_np, kernel_y, mode="nearest")
        slope_matrix_gpu = np.degrees(np.arctan(np.sqrt(dx**2 + dy**2)))
        aspect_matrix_gpu = 57.29578 * np.arctan2(dy, -dx)
        aspect_matrix_gpu = (aspect_matrix_gpu + 360) % 360
        print("全图坡度坡向计算完成")

    # ---------------- 计算每个工作表的地形特征 ----------------
    
    MATCH_TOLERANCE = 50.0  # 超过50米的匹配视为无效(写程序时示例数据有部分无法计算而且还影响程序，只好排除)
    results_dict = {}  # 存储所有结果

    for sheet_name, (sheet_x_np, sheet_y_np, _) in all_results.items():
        print(f"\n===== 计算工作表：{sheet_name} =====")
        num_total = len(sheet_x_np)
        print(f"原始目标点数量：{num_total}")

        # 跳过空数据
        if num_total == 0:
            results_dict[sheet_name] = []
            print("⚠️ 无有效目标点，跳过")
            continue

        # ---------------- 坐标匹配 ----------------
        if USE_GPU:
            # GPU模式：匹配坐标+计算距离
            target_x_gpu = jax.device_put(jnp.array(sheet_x_np), METAL_DEVICE)
            target_y_gpu = jax.device_put(jnp.array(sheet_y_np), METAL_DEVICE)
            
            # 调用增强版函数，获取索引和距离
            x_indices, x_distances = find_nearest_indices(x_coords_gpu, target_x_gpu)
            y_indices, y_distances = find_nearest_indices(y_coords_gpu, target_y_gpu)
            
            # 转回NumPy用于筛选
            x_indices = np.array(x_indices)
            y_indices = np.array(y_indices)
            x_distances = np.array(x_distances)
            y_distances = np.array(y_distances)
        else:
            # CPU模式：匹配坐标+计算距离
            x_dist_matrix = np.abs(x_coords_np[None, :] - sheet_x_np[:, None])
            x_indices = np.argmin(x_dist_matrix, axis=1)
            x_distances = np.min(x_dist_matrix, axis=1)
            
            y_dist_matrix = np.abs(y_coords_np[None, :] - sheet_y_np[:, None])
            y_indices = np.argmin(y_dist_matrix, axis=1)
            y_distances = np.min(y_dist_matrix, axis=1)

        # ---------------- 筛选有效匹配 ----------------
        # 仅保留x和y距离都≤阈值的点
        valid_mask = (x_distances <= MATCH_TOLERANCE) & (y_distances <= MATCH_TOLERANCE)
        valid_x_idx = x_indices[valid_mask]
        valid_y_idx = y_indices[valid_mask]
        valid_x_dist = x_distances[valid_mask]
        valid_y_dist = y_distances[valid_mask]
        valid_x = sheet_x_np[valid_mask]
        valid_y = sheet_y_np[valid_mask]
        num_valid = len(valid_x)

        # 打印筛选结果
        print(f"有效匹配点数量：{num_valid}（过滤{num_total - num_valid}个无效点）")
        if num_valid == 0:
            results_dict[sheet_name] = []
            print(f"⚠️ 所有点距离超过{MATCH_TOLERANCE}米，无有效匹配")
            continue
        else:
            print(f"有效点距离范围：x({valid_x_dist.min():.2f}~{valid_x_dist.max():.2f}m)，y({valid_y_dist.min():.2f}~{valid_y_dist.max():.2f}m)")

        # ---------------- 提取基础地形特征 ----------------
        # 高程、坡度、坡向
        elevations = np.array(e_matrix_gpu[valid_y_idx, valid_x_idx])
        slopes = np.array(slope_matrix_gpu[valid_y_idx, valid_x_idx])
        aspects = np.array(aspect_matrix_gpu[valid_y_idx, valid_x_idx])

        # ---------------- 计算3x3窗口统计特征 ----------------
        if USE_GPU:
            # GPU模式：批量计算窗口统计
            valid_y_gpu = jax.device_put(jnp.array(valid_y_idx), METAL_DEVICE)
            valid_x_gpu = jax.device_put(jnp.array(valid_x_idx), METAL_DEVICE)
            max_elev, min_elev, mean_elev, std_elev = window_stats_gpu(e_matrix_gpu, valid_y_gpu, valid_x_gpu)
            
            # 转回NumPy
            max_elev = np.array(max_elev)
            min_elev = np.array(min_elev)
            mean_elev = np.array(mean_elev)
            std_elev = np.array(std_elev)
        else:
            # CPU模式：计算窗口统计
            max_elev, min_elev, mean_elev, std_elev = [], [], [], []
            half_win = 1
            for y_idx, x_idx in zip(valid_y_idx, valid_x_idx):
                # 边界处理
                y0 = max(0, y_idx - half_win)
                y1 = min(e_matrix_np.shape[0], y_idx + half_win + 1)
                x0 = max(0, x_idx - half_win)
                x1 = min(e_matrix_np.shape[1], x_idx + half_win + 1)
                
                window = e_matrix_np[y0:y1, x0:x1]
                valid_window = window[~np.isnan(window)]
                
                if len(valid_window) >= 5:
                    max_elev.append(np.max(valid_window))
                    min_elev.append(np.min(valid_window))
                    mean_elev.append(np.mean(valid_window))
                    std_elev.append(np.std(valid_window))
                else:
                    # 数据不足时用中心值
                    center_val = e_matrix_np[y_idx, x_idx]
                    max_elev.append(center_val)
                    min_elev.append(center_val)
                    mean_elev.append(center_val)
                    std_elev.append(0.0)
            
            # 转为数组
            max_elev = np.array(max_elev)
            min_elev = np.array(min_elev)
            mean_elev = np.array(mean_elev)
            std_elev = np.array(std_elev)

        # ---------------- 计算衍生特征 ----------------
        elevation_range = max_elev - min_elev  # 高程极差
        rel_elev = elevations - mean_elev     # 相对高程

        # ---------------- 整理结果 ----------------
        sheet_result = []
        for idx in range(num_valid):
            feature = {
                "序号": idx + 1,
                "工作表名称": sheet_name,
                "原始x坐标/m": round(valid_x[idx], 3),
                "原始y坐标/m": round(valid_y[idx], 3),
                "匹配x距离/m": round(valid_x_dist[idx], 3),  # 新增：匹配距离，便于验证
                "匹配y距离/m": round(valid_y_dist[idx], 3),
                "高程/m": round(elevations[idx], 3),
                "坡度/°": round(slopes[idx], 3),
                "坡向/°": round(aspects[idx], 3),
                "局部最大高程/m": round(max_elev[idx], 3),
                "局部最小高程/m": round(min_elev[idx], 3),
                "局部平均高程/m": round(mean_elev[idx], 3),
                "高程极差/m": round(elevation_range[idx], 3),
                "地表粗糙度": round(std_elev[idx], 3),
                "相对高程/m": round(rel_elev[idx], 3)
            }
            sheet_result.append(feature)
            # 每100个点打印一次进度
            if (idx + 1) % 100 == 0 or (idx + 1) == num_valid:
                print(f"✅ 已处理{idx + 1}/{num_valid}个有效点")

        # 保存当前工作表结果
        results_dict[sheet_name] = sheet_result
        print(f"✅ 【{sheet_name}】计算完成，共{len(sheet_result)}条有效结果")

    # ---------------- 写入Excel并预览结果 ----------------
    write_multi_sheet_excel(results_dict, OUTPUT_EXCEL)

    # 结果预览
    print("\n===== 结果预览 =====")
    for sheet_name, result in results_dict.items():
        if result:
            print(f"\n【{sheet_name}】前3条结果：")
            preview_df = pd.DataFrame(result[:3])
            print(preview_df.to_string(index=False, max_colwidth=15))
        else:
            print(f"\n【{sheet_name}】无有效结果")

    print(f"\n===== 程序运行完成 =====")

# ===================== 运行主程序 =====================
if __name__ == "__main__":
    main()